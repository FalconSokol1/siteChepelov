import csv
import io
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.text import slugify

from .models import Category, ExtraService, Product, ProductImage

MAX_IMPORT_SIZE = 5 * 1024 * 1024
ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}

PRODUCT_ALIASES = {
    'артикул': 'sku',
    'код': 'sku',
    'название': 'name',
    'ссылка': 'slug',
    'материал': 'material',
    'стиль': 'style',
    'тип': 'product_type',
    'описание материала': 'material_label',
    'цена': 'price',
    'главное фото': 'image_url',
    'фото': 'images',
    'описание': 'description',
    'размеры': 'dimensions',
    'обработка': 'finish',
    'избранное': 'featured',
    'категория': 'category',
    'доп услуги': 'extra_services',
    'доп. услуги': 'extra_services',
}

SERVICE_ALIASES = {
    'код': 'slug',
    'название': 'title',
    'описание': 'description',
    'цена': 'price',
    'активна': 'is_active',
    'порядок': 'sort_order',
}


def _normalize_header(value, aliases):
    key = str(value or '').strip().lower()
    return aliases.get(key, key)


def _clean_row(headers, values, aliases):
    return {
        _normalize_header(header, aliases): value
        for header, value in zip(headers, values)
        if header not in (None, '')
    }


def _as_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_int(value, default=0, allow_none=False):
    text = _as_text(value).replace(' ', '').replace(',', '.')
    if not text:
        return None if allow_none else default
    try:
        return int(float(text))
    except (TypeError, ValueError) as exc:
        raise ValidationError(f'Некорректное число: {value}') from exc


def _as_bool(value, default=False):
    text = _as_text(value).lower()
    if not text:
        return default
    return text in {'1', 'true', 'yes', 'да', 'y', 'on'}


def _split(value):
    return [part.strip() for part in _as_text(value).split('|') if part.strip()]


def _validate_upload(upload):
    extension = Path(upload.name).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValidationError('Разрешены только файлы CSV и XLSX.')
    if upload.size > MAX_IMPORT_SIZE:
        raise ValidationError('Файл слишком большой. Максимальный размер — 5 МБ.')
    return extension


def _read_csv(upload):
    try:
        text = upload.read().decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise ValidationError('CSV должен быть сохранён в кодировке UTF-8.') from exc
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValidationError('Файл пуст.')
    return rows[0], rows[1:]


def _read_workbook(upload):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError('Для XLSX установите пакет openpyxl.') from exc

    workbook = load_workbook(upload, read_only=True, data_only=True)
    sheets = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            sheets[sheet.title.strip().lower()] = (rows[0], rows[1:])
    return sheets


def _import_services(headers, rows):
    created = updated = skipped = 0
    errors = []
    for row_number, values in enumerate(rows, start=2):
        row = _clean_row(headers, values, SERVICE_ALIASES)
        if not any(value not in (None, '') for value in row.values()):
            continue
        title = _as_text(row.get('title'))
        slug = _as_text(row.get('slug')) or slugify(title, allow_unicode=False)
        if not title or not slug:
            errors.append(f'Услуги, строка {row_number}: нужны код и название.')
            skipped += 1
            continue
        defaults = {
            'title': title,
            'description': _as_text(row.get('description')),
            'price': _as_int(row.get('price'), allow_none=True),
            'is_active': _as_bool(row.get('is_active'), True),
            'sort_order': _as_int(row.get('sort_order')),
        }
        _, was_created = ExtraService.objects.update_or_create(slug=slug, defaults=defaults)
        created += int(was_created)
        updated += int(not was_created)
    return created, updated, skipped, errors


def _import_products(headers, rows):
    created = updated = skipped = 0
    errors = []
    valid_materials = {key for key, _ in Product.MATERIAL_CHOICES}
    valid_styles = {key for key, _ in Product.STYLE_CHOICES}
    valid_types = {key for key, _ in Product.TYPE_CHOICES}

    for row_number, values in enumerate(rows, start=2):
        row = _clean_row(headers, values, PRODUCT_ALIASES)
        if not any(value not in (None, '') for value in row.values()):
            continue
        try:
            sku = _as_text(row.get('sku'))
            name = _as_text(row.get('name'))
            if not sku or not name:
                raise ValidationError('обязательны артикул и название')

            material = _as_text(row.get('material')) or 'granite'
            style = _as_text(row.get('style')) or 'classic'
            product_type = _as_text(row.get('product_type')) or 'single'
            if material not in valid_materials:
                raise ValidationError(f'неизвестный материал «{material}»')
            if style not in valid_styles:
                raise ValidationError(f'неизвестный стиль «{style}»')
            if product_type not in valid_types:
                raise ValidationError(f'неизвестный тип «{product_type}»')

            category = None
            category_value = _as_text(row.get('category'))
            if category_value:
                category = Category.objects.filter(slug=category_value).first()
                if not category:
                    category = Category.objects.filter(title__iexact=category_value).first()

            slug = _as_text(row.get('slug')) or slugify(name, allow_unicode=False) or sku.lower()
            image_urls = _split(row.get('images'))
            primary_image = _as_text(row.get('image_url')) or (image_urls[0] if image_urls else '')
            defaults = {
                'slug': slug,
                'name': name,
                'material': material,
                'style': style,
                'product_type': product_type,
                'material_label': _as_text(row.get('material_label')),
                'price': _as_int(row.get('price')),
                'image_url': primary_image,
                'description': _as_text(row.get('description')),
                'dimensions': _as_text(row.get('dimensions')),
                'finish': _as_text(row.get('finish')),
                'featured': _as_bool(row.get('featured')),
                'category': category,
            }
            product, was_created = Product.objects.update_or_create(sku=sku, defaults=defaults)

            if image_urls:
                product.images.all().delete()
                ProductImage.objects.bulk_create([
                    ProductImage(
                        product=product,
                        image_url=url,
                        alt_text=name,
                        sort_order=index,
                    )
                    for index, url in enumerate(image_urls)
                ])

            service_slugs = _split(row.get('extra_services'))
            if service_slugs:
                product.extra_services.set(
                    ExtraService.objects.filter(slug__in=service_slugs, is_active=True)
                )

            created += int(was_created)
            updated += int(not was_created)
        except ValidationError as exc:
            errors.append(f'Товары, строка {row_number}: {"; ".join(exc.messages)}.')
            skipped += 1
        except Exception as exc:
            errors.append(f'Товары, строка {row_number}: {exc}.')
            skipped += 1

    return created, updated, skipped, errors


@transaction.atomic
def import_catalog(upload):
    extension = _validate_upload(upload)
    if extension == '.csv':
        headers, rows = _read_csv(upload)
        product_result = _import_products(headers, rows)
        service_result = (0, 0, 0, [])
    else:
        sheets = _read_workbook(upload)
        product_sheet = sheets.get('products') or sheets.get('товары')
        service_sheet = sheets.get('services') or sheets.get('услуги')
        if not product_sheet and not service_sheet:
            raise ValidationError(
                'В XLSX нужен лист «products»/«товары» или «services»/«услуги».'
            )
        product_result = _import_products(*product_sheet) if product_sheet else (0, 0, 0, [])
        service_result = _import_services(*service_sheet) if service_sheet else (0, 0, 0, [])

    return {
        'products_created': product_result[0],
        'products_updated': product_result[1],
        'products_skipped': product_result[2],
        'services_created': service_result[0],
        'services_updated': service_result[1],
        'services_skipped': service_result[2],
        'errors': [*product_result[3], *service_result[3]],
    }
