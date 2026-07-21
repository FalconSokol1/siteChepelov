"""Выгрузка и загрузка прайс-листа Яндекс Бизнес (.xls / .xlsx) из живой БД."""

from __future__ import annotations

import io
import re
from pathlib import Path
from urllib.parse import urlparse

import xlwt
from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.text import slugify

from .models import Category, ExtraService, Product, ProductImage

MAX_IMPORT_SIZE = 5 * 1024 * 1024
ALLOWED_EXTENSIONS = {'.xls', '.xlsx'}
MIN_PRICE = 10_000

# Как в официальном шаблоне Яндекс Бизнес
YANDEX_HEADERS = [
    'Категория',
    'Наименование',
    'Идентификатор',
    'Описание',
    'Краткое описание',
    'Цена',
    'Фото',
    'Популярный товар',
    'В наличии',
    'Количество',
    'Единицы измерения',
    'Ссылка',
]

HEADER_ALIASES = {
    'категория': 'category',
    'наименование': 'name',
    'название': 'name',
    'идентификатор': 'sku',
    'артикул': 'sku',
    'описание': 'description',
    'краткое описание': 'short',
    'короткое описание': 'short',
    'цена': 'price',
    'фото': 'image',
    'популярный товар': 'popular',
    'в наличии': 'in_stock',
    'количество': 'qty',
    'единицы измерения': 'unit',
    'единица измерения': 'unit',
    'ссылка': 'url',
}

DEFAULT_IMAGE = (
    'https://lh3.googleusercontent.com/aida-public/'
    'AB6AXuB3TQzYLSdzTQdNcPY6wjwxodpJa75sDp0ERJGMEq6VSWhAaSauc8QElqTGJv1tM2IBG2fzfW4_'
    'R26FMZNSAdQW6kqSG3mtl8fORv0oRPQtVhODMtwykuZ4ywctIoO8G0rVx_QAaZR7OIsiGyaAiAJRC5__'
    'hbHWGIIjkcp0EmdMKV1XyAr3PQpiMN-WeYxQvO4xerRyZFeF80lDyupwaZh7ayfRbRpMCmZ8l7ViHcBm'
    'RER1uhEPzsjv4DAV9tozO4mIAGifKhqyYZc'
)


def site_base() -> str:
    return getattr(settings, 'SITE_URL', 'https://kavkazkamen.ru').rstrip('/')


def absolute_url(url: str) -> str:
    text = (url or '').strip()
    if not text:
        return DEFAULT_IMAGE
    if text.startswith('http://') or text.startswith('https://'):
        return text
    if text.startswith('/'):
        return f'{site_base()}{text}'
    return f'{site_base()}/{text}'


def _as_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_int(value, default=0) -> int:
    text = _as_text(value).replace(' ', '').replace(',', '.')
    if not text:
        return default
    try:
        return int(float(text))
    except (TypeError, ValueError) as exc:
        raise ValidationError(f'Некорректная цена или число: {value}') from exc


def _as_bool_popular(value) -> bool:
    text = _as_text(value).lower()
    return text in {'да', 'yes', 'true', '1', 'y'}


def _is_service_category(category: str) -> bool:
    text = category.lower()
    return 'услуг' in text


def _guess_material(name: str, short: str, description: str) -> str:
    blob = f'{name} {short} {description}'.lower()
    if 'мрамор' in blob:
        return 'marble'
    if 'лабрадор' in blob:
        return 'labradorite'
    if 'габбро' in blob:
        return 'gabbro'
    return 'granite'


def _guess_product_type(category: str, name: str) -> str:
    blob = f'{category} {name}'.lower()
    if 'сво' in blob:
        return 'svo'
    if 'двойн' in blob or 'дуэт' in blob:
        return 'double'
    if 'комплекс' in blob:
        return 'complex'
    return 'single'


def _guess_style(short: str, description: str) -> str:
    blob = f'{short} {description}'.lower()
    if 'минимал' in blob or 'современ' in blob:
        return 'modern'
    if 'автор' in blob:
        return 'author'
    return 'classic'


def _unique_slug(base: str, *, model, exclude_pk=None) -> str:
    root = slugify(base, allow_unicode=False) or 'item'
    root = root[:180]
    candidate = root
    index = 2
    while True:
        qs = model.objects.filter(slug=candidate)
        if exclude_pk:
            qs = qs.exclude(pk=exclude_pk)
        if not qs.exists():
            return candidate
        candidate = f'{root}-{index}'
        index += 1


def _slug_from_url(url: str) -> str:
    path = urlparse(_as_text(url)).path.strip('/')
    if not path:
        return ''
    part = path.split('/')[-1]
    return slugify(part, allow_unicode=False)


def product_to_row(product: Product) -> list:
    category_title = product.category.title if product.category_id else 'Памятники'
    short = product.material_label or product.dimensions or product.name
    return [
        category_title,
        product.name,
        product.sku,
        product.description or f'Премиальный памятник «{product.name}» из коллекции KavkazKamen.',
        short,
        max(int(product.price or 0), MIN_PRICE),
        absolute_url(product.image_url),
        'Да' if product.featured else '',
        'Да',
        1,
        'штука',
        f'{site_base()}/catalog/{product.slug}',
    ]


def service_to_row(service: ExtraService) -> list:
    price = max(int(service.price or 0), MIN_PRICE)
    return [
        'Дополнительные услуги',
        service.title,
        service.slug.upper() if service.slug else f'SRV-{service.pk}',
        service.description or service.title,
        service.title,
        price,
        absolute_url('/logo.png'),
        '',
        'Да',
        1,
        'штука',
        f'{site_base()}/catalog',
    ]


def build_rows(*, include_services: bool = True) -> list[list]:
    rows = [product_to_row(p) for p in Product.objects.select_related('category').all()]
    if include_services:
        rows.extend(
            service_to_row(s)
            for s in ExtraService.objects.filter(is_active=True).order_by('sort_order', 'title')
        )
    return rows


def export_xls_bytes(*, include_services: bool = True) -> bytes:
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Прайс-лист')

    for col, header in enumerate(YANDEX_HEADERS):
        ws.write(0, col, header)

    for row_idx, row in enumerate(build_rows(include_services=include_services), start=1):
        for col, value in enumerate(row):
            ws.write(row_idx, col, value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _normalize_headers(raw_headers: list) -> list[str]:
    mapped = []
    for header in raw_headers:
        key = _as_text(header).lower().replace('\xa0', ' ')
        mapped.append(HEADER_ALIASES.get(key, key))
    return mapped


def _read_xlsx_rows(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = _normalize_headers(list(next(rows_iter)))
    except StopIteration as exc:
        raise ValidationError('Файл пустой.') from exc

    result = []
    for values in rows_iter:
        if not values or all(v is None or _as_text(v) == '' for v in values):
            continue
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        result.append(row)
    return result


def _read_xls_rows(content: bytes) -> list[dict]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValidationError('Для чтения .xls установите пакет xlrd.') from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        # xlrd 2.x не читает .xls
        raise ValidationError(
            'Не удалось прочитать .xls. Сохраните файл как .xlsx или установите xlrd<2.'
        ) from exc

    sheet = book.sheet_by_index(0)
    if sheet.nrows < 2:
        raise ValidationError('В файле нет строк с товарами.')

    headers = _normalize_headers([sheet.cell_value(0, c) for c in range(sheet.ncols)])
    result = []
    for r in range(1, sheet.nrows):
        values = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        if all(_as_text(v) == '' for v in values):
            continue
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        result.append(row)
    return result


def read_price_rows(upload) -> list[dict]:
    name = getattr(upload, 'name', '') or ''
    extension = Path(name).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValidationError('Разрешены только файлы .xls и .xlsx')

    size = getattr(upload, 'size', None)
    if size is not None and size > MAX_IMPORT_SIZE:
        raise ValidationError('Файл слишком большой. Максимальный размер — 5 МБ.')

    content = upload.read()
    if extension == '.xlsx':
        return _read_xlsx_rows(content)
    return _read_xls_rows(content)


def _resolve_category(title: str) -> Category | None:
    text = _as_text(title)
    if not text:
        return None
    category = Category.objects.filter(title__iexact=text).first()
    if category:
        return category
    return Category.objects.filter(title__icontains=text).first()


def _import_service(row: dict) -> tuple[str, ExtraService]:
    name = _as_text(row.get('name'))
    sku = _as_text(row.get('sku')) or slugify(name, allow_unicode=False)
    if not name:
        raise ValidationError('У услуги нет наименования.')

    slug_base = slugify(sku, allow_unicode=False) or slugify(name, allow_unicode=False) or 'service'
    price = max(_as_int(row.get('price'), MIN_PRICE), MIN_PRICE)
    description = _as_text(row.get('description')) or _as_text(row.get('short'))

    existing = ExtraService.objects.filter(slug=slug_base).first()
    if not existing and sku:
        existing = ExtraService.objects.filter(slug__iexact=sku).first()

    defaults = {
        'title': name,
        'description': description,
        'price': price,
        'is_active': True,
    }
    if existing:
        for key, value in defaults.items():
            setattr(existing, key, value)
        existing.save()
        return 'updated', existing

    service = ExtraService.objects.create(
        slug=_unique_slug(slug_base, model=ExtraService),
        sort_order=ExtraService.objects.count(),
        **defaults,
    )
    return 'created', service


def _import_product(row: dict) -> tuple[str, Product]:
    name = _as_text(row.get('name'))
    sku = _as_text(row.get('sku'))
    if not name:
        raise ValidationError('У товара нет наименования.')
    if not sku:
        sku = f'KK-{slugify(name, allow_unicode=False)[:20].upper() or "ITEM"}'

    short = _as_text(row.get('short'))
    description = _as_text(row.get('description'))
    category_title = _as_text(row.get('category'))
    image = _as_text(row.get('image')) or DEFAULT_IMAGE
    price = max(_as_int(row.get('price'), MIN_PRICE), MIN_PRICE)
    featured = _as_bool_popular(row.get('popular'))
    url_slug = _slug_from_url(row.get('url'))
    material = _guess_material(name, short, description)
    style = _guess_style(short, description)
    product_type = _guess_product_type(category_title, name)
    material_label = short or name.upper()
    category = _resolve_category(category_title)

    existing = Product.objects.filter(sku=sku).first()
    slug = url_slug or (existing.slug if existing else '') or slugify(name, allow_unicode=False) or sku.lower()

    defaults = {
        'slug': _unique_slug(slug, model=Product, exclude_pk=existing.pk if existing else None),
        'name': name,
        'material': material,
        'style': style,
        'product_type': product_type,
        'material_label': material_label[:200],
        'price': price,
        'image_url': image[:1000],
        'description': description,
        'dimensions': short[:200] if re.search(r'\d', short) else (existing.dimensions if existing else ''),
        'finish': existing.finish if existing else '',
        'featured': featured,
        'category': category,
    }

    if existing:
        for key, value in defaults.items():
            setattr(existing, key, value)
        existing.save()
        product = existing
        created = False
    else:
        product = Product.objects.create(sku=sku[:32], **defaults)
        created = True

    if image and (created or not product.images.exists()):
        ProductImage.objects.filter(product=product).delete()
        ProductImage.objects.create(
            product=product,
            image_url=image[:1000],
            alt_text=name,
            sort_order=0,
        )

    return ('created' if created else 'updated'), product


@transaction.atomic
def import_price_list(upload) -> dict:
    rows = read_price_rows(upload)
    if not rows:
        raise ValidationError('В файле нет строк для импорта.')

    created_products = updated_products = 0
    created_services = updated_services = 0
    errors: list[str] = []

    for index, row in enumerate(rows, start=2):
        try:
            category = _as_text(row.get('category'))
            if _is_service_category(category):
                action, _ = _import_service(row)
                if action == 'created':
                    created_services += 1
                else:
                    updated_services += 1
            else:
                action, _ = _import_product(row)
                if action == 'created':
                    created_products += 1
                else:
                    updated_products += 1
        except ValidationError as exc:
            msg = '; '.join(getattr(exc, 'messages', []) or [str(exc)])
            errors.append(f'Строка {index}: {msg}')
        except Exception as exc:  # noqa: BLE001 — собираем ошибки по строкам
            errors.append(f'Строка {index}: {exc}')

    if errors and (created_products + updated_products + created_services + updated_services) == 0:
        raise ValidationError(errors[0] if len(errors) == 1 else 'Ошибки импорта: ' + '; '.join(errors[:5]))

    return {
        'created_products': created_products,
        'updated_products': updated_products,
        'created_services': created_services,
        'updated_services': updated_services,
        'errors': errors,
        'total_rows': len(rows),
    }
